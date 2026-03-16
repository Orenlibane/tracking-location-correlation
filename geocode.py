"""
Geocode all locations from the Excel file and save as JSON.
Uses Nominatim (OpenStreetMap) geocoder with SSL workaround.
"""
import json
import time
import ssl
import urllib.request
import urllib.parse
import openpyxl

EXCEL_PATH = '/Users/orentzezana/Downloads/FWEFEWF.xlsx'
OUTPUT_PATH = '/Users/orentzezana/Desktop/my-test-rpg-dungeon/map-tracking/locations.json'

# SSL context that doesn't verify certificates (needed on some macOS Python installs)
ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE

# Read Excel
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[wb.sheetnames[0]]

locations = []
seen = set()

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    _, b_val, c_val = row[0], row[1], row[2]
    if b_val and c_val:
        name = str(b_val).strip()
        area = str(c_val).strip()
        if name and area and name != 'כתובת' and name != 'אזור':
            if name not in seen:
                seen.add(name)
                locations.append({'name': name, 'area': area})

print(f"Total unique locations to geocode: {len(locations)}")

# Geocode using Nominatim
def geocode(name):
    query = urllib.parse.urlencode({
        'q': f'{name}, Israel',
        'format': 'json',
        'limit': 1,
        'countrycodes': 'il',
        'accept-language': 'he'
    })
    url = f'https://nominatim.openstreetmap.org/search?{query}'
    req = urllib.request.Request(url, headers={'User-Agent': 'MapTracker/1.0'})
    try:
        with urllib.request.urlopen(req, timeout=10, context=ssl_ctx) as resp:
            data = json.loads(resp.read().decode())
            if data:
                return float(data[0]['lat']), float(data[0]['lon'])
    except Exception as e:
        print(f"  Error geocoding {name}: {e}")
    return None, None

results = []
failed = []

for i, loc in enumerate(locations):
    lat, lon = geocode(loc['name'])
    time.sleep(1.1)  # Nominatim rate limit: 1 req/sec

    if lat is not None:
        results.append({
            'name': loc['name'],
            'area': loc['area'],
            'lat': lat,
            'lon': lon
        })
        print(f"  [{i+1}/{len(locations)}] {loc['name']} -> {lat:.4f}, {lon:.4f}")
    else:
        failed.append(loc)
        print(f"  [{i+1}/{len(locations)}] {loc['name']} -> FAILED")

# Save results
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump({'locations': results, 'failed': [f['name'] for f in failed]}, f, ensure_ascii=False, indent=2)

print(f"\nDone! {len(results)} geocoded, {len(failed)} failed.")
print(f"Saved to {OUTPUT_PATH}")
